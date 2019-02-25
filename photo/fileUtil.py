#encoding:utf-8:
import linecache

import openpyxl
from itertools import islice
import  csv

#txt 文件读写
# writeTxt(self，str) 续写tex，需要覆盖可以先执行flushTxt()清空
#flushTxt(self)清空txt文件
#readTxt(self，num),num>0读取，读取第num行，num<=0,读取所有行，包括空行。返回一个list
#redTxtNotNull(self)，读取所有行，不包括空行。返回一个list

class txtUtil(object):
    def __init__(self,path):
        self.path=path

    def writeTxt(self,str):
        f=open(self.path,'a')
        f.write(str)
        f.close()
    def flushTxt(self):
        f=open(self.path,'a')
        f.truncate()
        f.close()
    def readTxt(self,num):
        if num>0:
            line = linecache.getline(self.path,num)
            if line:
              return line
            else:
                return 0
        else:
            f=open(self.path,'r')
            list=[]
            for line in f:
               list.append(line.strip('\n'))
            f.close()
            return list
    def redTxtNotNull(self):
        f=open(self.path,'r')
        list=[]
        for line in f:
            s=line.strip('\n')
            if s:
                list.append(s)
        f.close()
        return list

#excel文件读写 基于openpyxl pip install openpyxl
#可用于后台简单导入导出
#readExcel逐行读，num为开始行数，end_为结束行数，返回一个list，list的元素是teple(),代表一列。readExcel(0,None)是读取所有行
#writeExcel逐行写，传入一个list，list的元素是teple(),代表一列
class excelUtil(object):
    def __init__(self,path):
        self.path=path
    def createExcel(self):
        wb=openpyxl.Workbook()
        wb.save(self.path)
    def readExcel(self,num,end_num):
        wb=openpyxl.load_workbook(self.path)
        sheet=wb.active
        list=[]
        for row in islice(sheet.values,num,end_num):
            list.append(row)
        return list
    def writeExcel(self,list):
        for x in list:
            wb=openpyxl.load_workbook(self.path)
            sheet=wb.active
            sheet.append(x)
            wb.save(self.path)

class csvlUtil(object):
    def __init__(self,path):
        self.path=path
    def redeCsv(self,num,end_num):
        pass


    def writeCsv(self,list):
        c = open(self.path,'ab')
        csv_write = csv.writer(c,dialect='excel')
        for x in list:
            csv_write.writerow(x)



if __name__=='__main__':
    t=excelUtil('10.xlsx')
    list=[(1,2,4,5),('欧杰文',4,5,6,7,8)]
    t.writeExcel(list)

